import csv
import sys
from openpyxl import load_workbook

students = []

def load_students_from_csv(file_name):
    """Load students from a CSV file."""
    global students
    try:
        with open(file_name, "r", newline="") as csvfile:
            reader = csv.DictReader(csvfile)
            students = list(reader)
            print(f"Loaded {len(students)} students from {file_name}.")
    except FileNotFoundError:
        print(f"Error: File {file_name} not found.")
    except Exception as e:
        print(f"Error loading file {file_name}: {e}")

def load_students_from_xlsx(file_name):
    """Load students from an XLSX file."""
    global students
    try:
        wb = load_workbook(filename=file_name)
        sheet = wb.active
        students = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            students.append({"name": row[0], "phone": row[1], "email": row[2], "group": row[3]})
        print(f"Loaded {len(students)} students from {file_name}.")
    except FileNotFoundError:
        print(f"Error: File {file_name} not found.")
    except Exception as e:
        print(f"Error loading file {file_name}: {e}")

def save_students_to_csv(file_name):
    """Save students to a CSV file."""
    try:
        with open(file_name, "w", newline="") as csvfile:
            fieldnames = ["name", "phone", "email", "group"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(students)
            print(f"Saved {len(students)} students to {file_name}.")
    except Exception as e:
        print(f"Error saving file {file_name}: {e}")

def print_all_students():
    for student in students:
        print(f"Name: {student['name']}, Phone: {student['phone']}, Email: {student['email']}, Group: {student['group']}")

def add_new_student():
    name = input("Enter student name: ")
    phone = input("Enter student phone: ")
    email = input("Enter student email: ")
    group = input("Enter student group: ")

    new_student = {"name": name, "phone": phone, "email": email, "group": group}
    insert_position = 0
    for student in students:
        if new_student["name"] > student["name"]:
            insert_position += 1
        else:
            break
    students.insert(insert_position, new_student)
    print("Student has been added successfully.")

def delete_student():
    name = input("Enter the name of the student to delete: ")
    for student in students:
        if student["name"] == name:
            students.remove(student)
            print("Student has been deleted.")
            return
    print("Student not found.")

def update_student():
    name = input("Enter the name of the student to update: ")
    update_student = None
    for student in students:
        if student["name"] == name:
            update_student = student
            break

    if not update_student:
        print("Student not found.")
        return

    print(f"Current details - Name: {update_student['name']}, Phone: {update_student['phone']}, Email: {update_student['email']}, Group: {update_student['group']}")
    new_name = input(f"Enter new name (leave empty to keep '{update_student['name']}'): ") or update_student['name']
    new_phone = input(f"Enter new phone (leave empty to keep '{update_student['phone']}'): ") or update_student['phone']
    new_email = input(f"Enter new email (leave empty to keep '{update_student['email']}'): ") or update_student['email']
    new_group = input(f"Enter new group (leave empty to keep '{update_student['group']}'): ") or update_student['group']

    updated_student = {"name": new_name, "phone": new_phone, "email": new_email, "group": new_group}
    students.remove(update_student)

    insert_position = 0
    for student in students:
        if updated_student["name"] > student["name"]:
            insert_position += 1
        else:
            break

    students.insert(insert_position, updated_student)
    print("Student information has been updated successfully.")

def main():
    if len(sys.argv) < 3:
        print("Usage: python script.py <input_file> <file_type>")
        return

    input_file = sys.argv[1]
    file_type = sys.argv[2].lower()

    if file_type == "csv":
        load_students_from_csv(input_file)
    elif file_type == "xlsx":
        load_students_from_xlsx(input_file)
    else:
        print("Unsupported file type. Use 'csv' or 'xlsx'.")
        return

    while True:
        choice = input("Choose an action: [C]reate, [U]pdate, [D]elete, [P]rint, e[X]it: ").lower()
        if choice == 'c':
            add_new_student()
        elif choice == 'u':
            update_student()
        elif choice == 'd':
            delete_student()
        elif choice == 'p':
            print_all_students()
        elif choice == 'x':
            output_file = input("Enter the output file name to save changes (leave empty to skip): ")
            if output_file:
                save_students_to_csv(output_file)
            print("Exiting program.")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
